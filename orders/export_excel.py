"""
Выгрузка очереди заказов в Excel.
"""
import json

from django.http import HttpResponse
from django.utils import timezone

from .models import OrderSyncQueue


HEADERS = [
    "id", "action", "order_uuid", "created_at",
    "number", "user_uuid", "name", "email", "phone",
    "delivery_method", "delivery_city", "delivery_address",
    "payment_type", "total", "total_pv", "status", "comment", "created_at", "items",
]
PAYLOAD_KEYS = [
    "number", "user_uuid", "name", "email", "phone",
    "delivery_method", "delivery_city", "delivery_address",
    "payment_type", "total", "total_pv", "status", "comment", "created_at", "items",
]


def export_pending_to_excel(request):
    """Экспорт всех pending записей в Excel, затем пометка как sent."""
    from openpyxl import Workbook
    from openpyxl.styles import Font
    from openpyxl.utils import get_column_letter

    qs = OrderSyncQueue.objects.filter(status=OrderSyncQueue.Status.PENDING).order_by("created_at")
    rows = list(qs)

    wb = Workbook()
    ws = wb.active
    ws.title = "Очередь выгрузки заказов"

    ws.append(HEADERS)
    for cell in ws[1]:
        cell.font = Font(bold=True)

    for r in rows:
        payload = r.payload or {}
        row = [
            r.id,
            r.action,
            str(r.order_uuid) if r.order_uuid else "",
            r.created_at.strftime("%Y-%m-%d %H:%M:%S") if r.created_at else "",
        ]
        for k in PAYLOAD_KEYS:
            v = payload.get(k)
            if k == "items" and isinstance(v, list):
                row.append(json.dumps(v, ensure_ascii=False) if v else "")
            else:
                row.append(v if v is not None else "")
        ws.append(row)

    widths = [8, 12, 38, 18, 10, 38, 20, 28, 16, 16, 30, 30, 12, 12, 12, 12, 20, 22, 50]
    for i, w in enumerate(widths[: len(HEADERS)], 1):
        ws.column_dimensions[get_column_letter(i)].width = min(w, 50)

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    fname = f"order_sync_queue_{timezone.now().strftime('%Y%m%d_%H%M')}.xlsx"
    response["Content-Disposition"] = f'attachment; filename="{fname}"'

    wb.save(response)

    now = timezone.now()
    qs.update(status=OrderSyncQueue.Status.SENT, sent_at=now)

    return response
