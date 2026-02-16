"""
Выгрузка очереди пользователей в Excel.
"""
from django.http import HttpResponse
from django.utils import timezone

from .models import UserSyncQueue


HEADERS = ["id", "action", "user_uuid", "created_at", "uuid", "email", "first_name", "last_name", "phone", "is_business_user", "referred_by_uuid", "is_active", "date_joined"]
PAYLOAD_KEYS = ["uuid", "email", "first_name", "last_name", "phone", "is_business_user", "referred_by_uuid", "is_active", "date_joined"]


def export_pending_to_excel(request):
    """Экспорт всех pending записей в Excel, затем пометка как sent."""
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font
    from openpyxl.utils import get_column_letter

    qs = UserSyncQueue.objects.filter(status=UserSyncQueue.Status.PENDING).order_by("created_at")
    rows = list(qs)

    wb = Workbook()
    ws = wb.active
    ws.title = "Очередь выгрузки"

    ws.append(HEADERS)
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(wrap_text=True, vertical="top")

    for r in rows:
        payload = r.payload or {}
        row = [
            r.id,
            r.action,
            str(r.user_uuid) if r.user_uuid else "",
            r.created_at.strftime("%Y-%m-%d %H:%M:%S") if r.created_at else "",
        ]
        for k in PAYLOAD_KEYS:
            v = payload.get(k)
            row.append(v if v is not None else "")
        ws.append(row)

    # Ширина колонок
    widths = [8, 12, 38, 18, 38, 28, 18, 18, 16, 8, 38, 6, 22]
    for i, w in enumerate(widths[: len(HEADERS)], 1):
        ws.column_dimensions[get_column_letter(i)].width = min(w, 50)

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    fname = f"user_sync_queue_{timezone.now().strftime('%Y%m%d_%H%M')}.xlsx"
    response["Content-Disposition"] = f'attachment; filename="{fname}"'

    wb.save(response)

    now = timezone.now()
    qs.update(status=UserSyncQueue.Status.SENT, sent_at=now)

    return response
